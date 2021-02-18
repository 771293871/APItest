import requests
import time
import xlrd
import openpyxl
import pytest
import json
import ast
import time
from requests_toolbelt import MultipartEncoder

# casename=input('输入要执行的用例集名称：')

def strtodict(text1):  #将字符类型转为字典类型
    responsetext2 = ast.literal_eval(text1)
    return responsetext2
def get_all_case(casepath,sheetname):   #将用例集中的每行用例生成一个字典存放到列表中，返回列表。参数为存放用例的excel文档路径和用例集名称
    ex1=openpyxl.load_workbook(casepath)
    sh1=ex1.get_sheet_by_name(sheetname)
    rowmax=sh1.max_row    #最大行数
    colsmax=sh1.max_column   #最大列数
    header=[]
    for i in range(1, colsmax + 1):
        cellvalue = sh1.cell(row=1, column=i).value
        header.append(cellvalue)
    j=2
    data_list=[]
    while j<=rowmax:
        rowdata = []
        for i in range(1, colsmax + 1):
            cellvalue = sh1.cell(row=j, column=i).value
            # cellvalue=cellvalue.replace('\n','',1)
            rowdata.append(cellvalue)
        j=j+1
        d = dict(zip(header,rowdata))
        data_list.append(d)
    return data_list


def check_user_info(all_case):
    i=0
    j=len(all_case)
    while i<j:
        case1=all_case[i]
        # print(case1)
        response1 = requests.post(url=case1["url"], json=strtodict(case1["data"]),headers={'token':token1})
        responseinfo = strtodict(response1.text)
        case3 = case1['expect_res'].split(",")
        resmobile1=case3[0]
        resmobilenickname1=case3[1]
        a=0
        b=len(responseinfo['body']['data']['datas'])
        while a<b:
            resmobile2 = responseinfo['body']['data']['datas'][a]['mobile']
            resmobilenickname2 = responseinfo['body']['data']['datas'][a]['nickName']
            restotal2 = responseinfo['body']['data']['total']
            boola = resmobile2 == resmobile1
            boolb = resmobilenickname2 == resmobilenickname1
            boolc = restotal2 == b
            if boola == True and boolb == True  and boolc == True:
                break
            elif boola == False or boolb == False  or boolc == False:
                time.sleep(1)
            a=a+1
        if boola==True and boolb==True and boolc==True:
            print(case1['case_name'],':执行通过')
        elif boola==False or boolb==False  or boolc==False:
            print(boola,boolb,boolc)
            print(case1['case_name'],':查询结果的手机全号，昵称和查询结果总数失败')
        i=i+1


def casenameget_one_case(all_case,casename):   #按照测试用例名称获取对应的测试用例
    for case_data in all_case:
        if casename == case_data['case_name']:  # 如果字典数据中case_name与参数一致
            case1=case_data
    return case1


allcase=get_all_case('E:/pythonwork/APItest/testcase/testcase.xlsx','登录')
case2=casenameget_one_case(allcase,'user_login')
response1=requests.post(url=case2["url"],json=strtodict(case2["data"]))
responseinfo=strtodict(response1.text)
token1=responseinfo['token']
boola=responseinfo['msg']==case2["expect_res"]
if boola==True:
    print('登录执行通过')
elif boola==False:
    print('登录执行不通过')
check_user_info(get_all_case('E:/pythonwork/APItest/testcase/testcase.xlsx','用户查询'))








# allcase=get_all_case('E:/pythonwork/APItest/testcase/testcase.xlsx','用户查询')
#
# case2=get_one_case(allcase,'user_login')
# response1=requests.post(url=case2["url"],json=strtodict(case2["data"]))
# responseinfo=strtodict(response1.text)
# token1=responseinfo['token']
# boola=responseinfo['msg']==case2["expect_res"]
# if boola==True:
#     print('用例user_login执行通过')
# elif boola==False:
#     print('用例user_login执行不通过')



#
# case2=get_one_case(allcase,'mobile_jingquecheck_user')
# case3=case2['expect_res'].split(",")
# resmobile1=case3[0]
# resmobilenickname1=case3[1]
# restotal1=case3[2]
# response1=requests.post(url=case2["url"],json=strtodict(case2["data"]),headers={'token':token1})
# responseinfo=strtodict(response1.text)
# resmobile2=responseinfo['body']['data']['datas'][0]['mobile']
# resmobilenickname2=responseinfo['body']['data']['datas'][0]['nickName']
# restotal2=responseinfo['body']['data']['total']
# restotala2=responseinfo['body']
# boola=resmobile2==resmobile1
# boolb=resmobilenickname2==resmobilenickname1
# boolc=restotal2==restotal1
# if boola==True and boolb==True and boolc==True:
#     print('使用手机全号搜索用户，jingquemobile_check_user执行通过')
# elif boola==False or boolb==False or boolc==False:
#     print(boola,boolb,boolc)
#     print('使用手机全号搜索用户，jingquemobile_check_user执行不通过')


# case2=get_one_case(allcase,'emobile_mohucheck_user')
# case3=case2['expect_res'].split(",")
# resmobile1=case3[0]
# resmobilenickname1=case3[1]
# restotal1=case3[2]
# response1=requests.post(url=case2["url"],json=strtodict(case2["data"]),headers={'token':token1})
# print(response1.text)